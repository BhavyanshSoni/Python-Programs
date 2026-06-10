"""Export inventory and financial data to CSV and Excel."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from config import EXPORT_DIR
from database import Database
from analytics import calculate_dashboard_metrics


def _ensure_export_dir() -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    return EXPORT_DIR


def export_inventory_csv(db: Database, owner_id: int) -> Path:
    """Export current inventory snapshot to CSV."""
    export_dir = _ensure_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = export_dir / f"inventory_{timestamp}.csv"

    products = db.get_products(owner_id)
    with filepath.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "product_id", "name", "category", "cost_price_inr",
                "selling_price_inr", "quantity", "stock_value_cost_inr", "stock_value_retail_inr",
            ],
        )
        writer.writeheader()
        for p in products:
            writer.writerow(
                {
                    "product_id": p["product_id"],
                    "name": p["name"],
                    "category": p["category"],
                    "cost_price_inr": p["cost_price"],
                    "selling_price_inr": p["selling_price"],
                    "quantity": p["quantity"],
                    "stock_value_cost_inr": round(p["cost_price"] * p["quantity"], 2),
                    "stock_value_retail_inr": round(p["selling_price"] * p["quantity"], 2),
                }
            )
    return filepath


def export_financial_csv(db: Database, owner_id: int) -> Path:
    """Export sales transactions and summary metrics to CSV."""
    export_dir = _ensure_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = export_dir / f"financial_report_{timestamp}.csv"

    sales = db.get_recent_sales(owner_id, limit=100_000)
    metrics = calculate_dashboard_metrics(db, owner_id)

    with filepath.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Financial Summary (INR)"])
        writer.writerow(["Metric", "Value"])
        for key, value in metrics.items():
            writer.writerow([key.replace("_", " ").title(), value])
        writer.writerow([])
        writer.writerow(["Sales Transactions"])
        writer.writerow(
            ["ID", "SKU", "Product", "Qty", "Revenue (INR)", "Profit (INR)", "Sold At", "Sold By"]
        )
        for s in sales:
            writer.writerow(
                [
                    s["id"], s["product_id"], s["name"], s["quantity_sold"],
                    s["total_revenue"], s["profit"], s["sold_at"], s["sold_by"],
                ]
            )
    return filepath


def export_inventory_excel(db: Database, owner_id: int) -> Path:
    """Export inventory to Excel with formatted sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl") from exc

    export_dir = _ensure_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = export_dir / f"inventory_report_{timestamp}.xlsx"

    wb = Workbook()
    ws_inv = wb.active
    ws_inv.title = "Inventory"

    header_fill = PatternFill(start_color="1C2430", end_color="1C2430", fill_type="solid")
    header_font = Font(bold=True, color="F0F4F8")

    headers = ["SKU", "Name", "Category", "Cost (₹)", "Price (₹)", "Qty", "Value Cost (₹)", "Value Retail (₹)"]
    ws_inv.append(headers)
    for cell in ws_inv[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for p in db.get_products(owner_id):
        ws_inv.append(
            [
                p["product_id"], p["name"], p["category"],
                p["cost_price"], p["selling_price"], p["quantity"],
                round(p["cost_price"] * p["quantity"], 2),
                round(p["selling_price"] * p["quantity"], 2),
            ]
        )

    ws_fin = wb.create_sheet("Financial Summary")
    metrics = calculate_dashboard_metrics(db, owner_id)
    ws_fin.append(["Metric", "Value (INR)"])
    for key, value in metrics.items():
        ws_fin.append([key.replace("_", " ").title(), value])

    ws_sales = wb.create_sheet("Sales")
    ws_sales.append(["ID", "SKU", "Product", "Qty", "Revenue (₹)", "Profit (₹)", "Sold At", "Sold By"])
    for s in db.get_recent_sales(owner_id, limit=100_000):
        ws_sales.append(
            [
                s["id"], s["product_id"], s["name"], s["quantity_sold"],
                s["total_revenue"], s["profit"], s["sold_at"], s["sold_by"],
            ]
        )

    wb.save(filepath)
    return filepath
